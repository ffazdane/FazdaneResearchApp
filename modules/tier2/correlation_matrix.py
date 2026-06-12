"""
FazDane Analytics - Tier 2
Correlation Matrix, Asset Segmentation, and Down-Time Backtesting Module
========================================================================
Features:
- Live yfinance data fetching with automatic reference benchmark injection (SPY, QQQ, IWM).
- SQLite persistence of every run to data/correlation_analysis/correlation_analysis.sqlite.
- Immediate database run loading for instant navigation.
- Tab 1: Dynamic interactive correlation heatmap and styled data tables.
- Tab 2: Asset segmentation playbook, priority candidate list sub-tabs (Bull, Selloff, Chop), and master ranking.
- Tab 3: Index down-time backtest, high-level metrics, and Plotly performance charts.
- Styled dynamic multi-tab Excel exporter matching the house design system.
"""

import io
import math
from datetime import datetime
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import yfinance as yf
from pandas.io.formats.style import Styler

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from modules.base_module import FazDaneModule
from utils.universe_manager import get_universe_names, render_universe_manager
from modules.tier2 import correlation_store


DEFAULT_RENAME_MAP = {
    "SPY": "S&P 500",
    "QQQ": "NASDAQ",
    "IWM": "RUSSELL",
    "^VIX": "VOLATILITY",
    "TLT": "BONDS",
    "GLD": "GOLD",
    "CL=F": "OIL",
    "BTC=F": "BITCOIN",
    "DX-Y.NYB": "USD",
    "HG=F": "COPPER",
}


@st.cache_data(ttl=3600, show_spinner=False)
def fetch_close_prices(tickers: tuple[str, ...], start_date, end_date) -> tuple[pd.DataFrame, dict]:
    prices = pd.DataFrame()
    info = {}
    
    from utils.formatting import calculate_strength_pct, format_strength_meter
    from modules.trade_recommendation.indicators import calculate_fdts
    
    for symbol in tickers:
        try:
            data = yf.download(
                symbol,
                start=start_date,
                end=end_date,
                auto_adjust=True,
                progress=False,
            )
            if data.empty:
                continue

            if isinstance(data.columns, pd.MultiIndex):
                data.columns = data.columns.get_level_values(0)

            close = data["Close"] if "Close" in data.columns else data.iloc[:, 0]
            if isinstance(close, pd.DataFrame):
                close = close.iloc[:, 0]
            prices[symbol] = close
            
            # Strength calculation
            strength_pct = calculate_strength_pct(data)
            tri, _ = format_strength_meter(strength_pct)
            strength_str = f"{tri} {strength_pct:.1%}" if strength_pct is not None else "—"
            
            # FDTS calculation
            _, fdts_signals = calculate_fdts(data)
            fdts_val = str(fdts_signals.iloc[-1]) if len(fdts_signals) > 0 else "Neutral"
            fdts_emoji = {"Buy": "🟢 Buy", "Sell": "🔴 Sell", "Neutral": "⚪ Neutral"}.get(fdts_val, f"⚪ {fdts_val}")
            
            info[symbol] = {
                "Strength": strength_str,
                "FDTS": fdts_emoji
            }
        except Exception:
            continue

    return prices.dropna(how="all"), info


def compute_correlation(prices: pd.DataFrame, method: str) -> pd.DataFrame:
    returns = prices.ffill().pct_change().dropna(how="all")
    return returns.corr(method=method)


def format_correlation_table(corr: pd.DataFrame) -> Styler:
    table = corr.copy()
    table.index.name = "Symbol"
    return (
        table.style.format("{:.0%}")
        .set_properties(
            **{
                "color": "#ffffff",
                "font-size": "15px",
                "font-weight": "700",
                "text-align": "center",
            }
        )
        .set_table_styles(
            [
                {
                    "selector": "th",
                    "props": [
                        ("background-color", "#1f6f8b"),
                        ("color", "#ffffff"),
                        ("font-size", "15px"),
                        ("font-weight", "700"),
                        ("text-align", "center"),
                    ],
                }
            ]
        )
    )

# --------------------------------------------------------------------------- #
# Styler Color Helper Callbacks
# --------------------------------------------------------------------------- #

def color_correlation(val):
    if not isinstance(val, (int, float)) or pd.isna(val):
        return ""
    if val >= 0.5:
        return "background-color: #63BE7B; color: #111827; font-weight: bold;"
    elif val >= 0.25:
        return "background-color: #C6E0B4; color: #111827;"
    elif val <= -0.3:
        return "background-color: #F8696B; color: #ffffff; font-weight: bold;"
    elif val <= -0.1:
        return "background-color: #FCE4D6; color: #111827;"
    return ""


def color_return(val):
    if not isinstance(val, (int, float)) or pd.isna(val):
        return ""
    if val >= 1.0:
        return "background-color: #C6E0B4; color: #111827; font-weight: bold;"
    elif val > 0:
        return "background-color: #E2EFDA; color: #111827;"
    elif val <= -1.0:
        return "background-color: #F8696B; color: #ffffff; font-weight: bold;"
    elif val < 0:
        return "background-color: #FCE4D6; color: #111827;"
    return ""


def color_bucket(val):
    colors = {
        "1 - High-Beta Leaders": "background-color: #C6E0B4; color: #111827; font-weight: bold;",
        "2 - Core Trend": "background-color: #E2EFDA; color: #111827;",
        "3 - Low-Beta Followers": "background-color: #FFF2CC; color: #111827;",
        "4 - Decoupled / Neutral": "background-color: #DEEAF1; color: #111827;",
        "5 - Uncorrelated Hedge": "background-color: #D9D9D9; color: #111827;",
        "6 - Inverse (Sell-off Winners)": "background-color: #FCE4D6; color: #111827; font-weight: bold;",
    }
    return colors.get(val, "")


def style_fdts(val):
    if not isinstance(val, str):
        return ""
    if "Buy" in val:
        return "color: #22c55e; font-weight: 700;"
    elif "Sell" in val:
        return "color: #ef4444; font-weight: 700;"
    return ""


def style_strength(val):
    if not isinstance(val, str):
        return ""
    if "▲" in val:
        return "color: #00D4AA; font-weight: 700;"
    elif "▼" in val:
        return "color: #FF4B4B; font-weight: 700;"
    elif "▶" in val:
        return "color: #FFA421; font-weight: 700;"
    return ""

# --------------------------------------------------------------------------- #
# Engine Computations
# --------------------------------------------------------------------------- #

def compute_all_results(prices: pd.DataFrame, tickers: list[str], method: str, info_dict: dict) -> dict:
    """Run full asset segmentation and down-time backtesting calculations."""
    rets = prices.ffill().pct_change().dropna(how="all") * 100.0
    corr = rets.corr(method=method)
    
    benchmarks = ["SPY", "QQQ", "IWM"]
    for b in benchmarks:
        if b not in corr.columns:
            corr[b] = 0.0
            corr.loc[b] = 0.0
            
    REFERENCE_TICKERS = {
        "DIA", "SPMO", "SMH", "BONDS", "ARKK", "EEM", "FXI", "GDX", "GDXJ", "SLV",
        "USO", "UNG", "XOP", "KRE",
    }
    
    PLAYBOOK = {
        "1 - High-Beta Leaders": {
            "Bull":    "PRIMARY - Call calendars (ride trend); slightly OTM call strikes",
            "Selloff": "AVOID / Put calendars - these fall hardest in risk-off",
            "Chop":    "ATM calendars - high theta, watch IV crush",
        },
        "2 - Core Trend": {
            "Bull":    "Core - Call calendars at-the-money to slightly OTM",
            "Selloff": "Reduce / Put calendars",
            "Chop":    "ATM calendars",
        },
        "3 - Low-Beta Followers": {
            "Bull":    "Secondary - milder trend participation",
            "Selloff": "Hold - smaller drawdown than leaders",
            "Chop":    "Good ATM calendar candidates (lower whipsaw)",
        },
        "4 - Decoupled / Neutral": {
            "Bull":    "Neutral - stock-specific, size down",
            "Selloff": "Neutral - relative safe harbor",
            "Chop":    "PRIMARY - delta-neutral ATM calendars",
        },
        "5 - Uncorrelated Hedge": {
            "Bull":    "Diversifier - low trend capture",
            "Selloff": "Defensive ballast",
            "Chop":    "Neutral calendars / portfolio ballast",
        },
        "6 - Inverse (Sell-off Winners)": {
            "Bull":    "AVOID / Put calendars - tend to fall when market rises",
            "Selloff": "PRIMARY - Call calendars (these RISE in risk-off)",
            "Chop":    "Tactical hedge overlay",
        },
    }
    
    display_tickers = [t for t in tickers if t in corr.columns]
    display_corr = corr.loc[display_tickers, display_tickers]
    
    rows = []
    valid_tickers = [t for t in tickers if t in corr.columns and t not in benchmarks]
    info_dict = info_dict or {}
    
    for t in valid_tickers:
        c_spy = float(corr.loc[t, "SPY"])
        c_qqq = float(corr.loc[t, "QQQ"])
        c_iwm = float(corr.loc[t, "IWM"])
        
        triplet = np.array([c_spy, c_qqq, c_iwm], dtype=float)
        composite = float(np.mean(triplet))
        min_c = float(np.min(triplet))
        max_c = float(np.max(triplet))
        spread = max_c - min_c
        
        best_bench = "SPY"
        if c_qqq > c_spy and c_qqq > c_iwm:
            best_bench = "QQQ"
        elif c_iwm > c_spy and c_iwm > c_qqq:
            best_bench = "IWM"
            
        row_vals = display_corr.loc[t].drop(labels=[t], errors="ignore").astype(float)
        avg_uni = float(row_vals.mean()) if len(row_vals) > 0 else 0.0
        idiosyncratic = 1.0 - avg_uni
        
        same_sign = (triplet > 0).all() or (triplet < 0).all()
        consistent = bool(same_sign and spread <= 0.20)
        
        realized_vol_20d = np.nan
        if t in rets.columns:
            asset_rets = rets[t].dropna()
            if len(asset_rets) >= 20:
                realized_vol_20d = float(asset_rets.iloc[-20:].std() * np.sqrt(252))
            elif len(asset_rets) > 1:
                realized_vol_20d = float(asset_rets.std() * np.sqrt(252))
                
        bull_priority = float(max(composite, 0.0) * 100.0 + int(consistent) * 8.0 - spread * 10.0)
        selloff_priority = float(max(-composite, 0.0) * 100.0 + int(consistent) * 8.0 - spread * 10.0)
        chop_priority = float((1.0 - abs(composite)) * 60.0 + idiosyncratic * 40.0)
        
        def get_bucket(c):
            if c >= 0.50:
                return "1 - High-Beta Leaders"
            if c >= 0.35:
                return "2 - Core Trend"
            if c >= 0.20:
                return "3 - Low-Beta Followers"
            if c >= 0.05:
                return "4 - Decoupled / Neutral"
            if c > -0.10:
                return "5 - Uncorrelated Hedge"
            return "6 - Inverse (Sell-off Winners)"
            
        bucket = get_bucket(composite)
        
        ticker_info = info_dict.get(t, {"Strength": "—", "FDTS": "⚪ Neutral"})
        strength_val = ticker_info.get("Strength", "—")
        fdts_val = ticker_info.get("FDTS", "⚪ Neutral")
        
        rows.append({
            "Ticker": t,
            "Is_Reference": t in REFERENCE_TICKERS,
            "Bucket": bucket,
            "Best_Proxy": best_bench,
            "FDTS": fdts_val,
            "Strength": strength_val,
            "Corr_SPY": round(c_spy, 4),
            "Corr_QQQ": round(c_qqq, 4),
            "Corr_IWM": round(c_iwm, 4),
            "Composite": round(composite, 4),
            "Min_Corr": round(min_c, 4),
            "Max_Corr": round(max_c, 4),
            "Bench_Spread": round(spread, 4),
            "Avg_Corr_Universe": round(avg_uni, 4),
            "Idiosyncratic_Score": round(idiosyncratic, 4),
            "Consistent": consistent,
            "RealizedVol_20d": round(realized_vol_20d, 2) if not np.isnan(realized_vol_20d) else None,
            "Bull_Priority": round(bull_priority, 2),
            "Selloff_Priority": round(selloff_priority, 2),
            "Chop_Priority": round(chop_priority, 2),
            "Bull_Action": PLAYBOOK[bucket]["Bull"],
            "Selloff_Action": PLAYBOOK[bucket]["Selloff"],
            "Chop_Action": PLAYBOOK[bucket]["Chop"],
        })
        
    segmentation_df = pd.DataFrame(rows)
    if not segmentation_df.empty:
        segmentation_df = segmentation_df.sort_values("Composite", ascending=False).reset_index(drop=True)
        segmentation_df.insert(0, "Rank", range(1, len(segmentation_df) + 1))
    else:
        segmentation_df = pd.DataFrame(columns=["Rank", "Ticker", "Is_Reference", "Bucket"])
        
    if not segmentation_df.empty:
        bucket_summary = (segmentation_df.groupby("Bucket")
                   .agg(Count=("Ticker", "size"),
                        Avg_Composite=("Composite", "mean"),
                        Avg_SPY=("Corr_SPY", "mean"),
                        Avg_QQQ=("Corr_QQQ", "mean"),
                        Avg_IWM=("Corr_IWM", "mean"),
                        Consistent_Names=("Consistent", "sum"))
                   .reset_index().sort_values("Bucket"))
        for c in ["Avg_Composite", "Avg_SPY", "Avg_QQQ", "Avg_IWM"]:
            bucket_summary[c] = bucket_summary[c].round(3)
    else:
        bucket_summary = pd.DataFrame(columns=["Bucket", "Count", "Avg_Composite"])
        
    backtest_tickers = [t for t in valid_tickers if t in rets.columns]
    bench_cols = ["SPY", "QQQ", "IWM"]
    for b in bench_cols:
        if b not in rets.columns:
            rets[b] = 0.0
            
    mkt_avg = rets[bench_cols].mean(axis=1)
    all_down = (rets["SPY"] < 0) & (rets["QQQ"] < 0) & (rets["IWM"] < 0)
    selloff_idx = rets.index[all_down]
    
    backtest_summary = None
    backtest_detail = None
    
    if len(selloff_idx) > 0:
        detail = pd.DataFrame(index=selloff_idx)
        detail["Date"] = selloff_idx.strftime("%Y-%m-%d")
        detail["SPY_%"] = rets.loc[selloff_idx, "SPY"].round(2)
        detail["QQQ_%"] = rets.loc[selloff_idx, "QQQ"].round(2)
        detail["IWM_%"] = rets.loc[selloff_idx, "IWM"].round(2)
        detail["Mkt_Avg_%"] = mkt_avg.loc[selloff_idx].round(2)
        
        def severity(x):
            if x <= -2.0:
                return "Severe (<=-2%)"
            if x <= -1.0:
                return "Major (-1 to -2%)"
            return "Mild (0 to -1%)"
            
        detail["Severity"] = detail["Mkt_Avg_%"].apply(severity)
        
        for t in backtest_tickers:
            detail[t] = rets.loc[selloff_idx, t].round(2)
            
        detail["Tickers_Up"] = (detail[backtest_tickers] > 0).sum(axis=1)
        detail["Pct_Tickers_Up"] = (detail["Tickers_Up"] / len(backtest_tickers) * 100.0).round(1) if len(backtest_tickers) > 0 else 0.0
        backtest_detail = detail.sort_index(ascending=False).reset_index(drop=True)
        
        sub_rets = rets.loc[selloff_idx]
        severe_mask = mkt_avg.loc[selloff_idx] <= -1.0
        
        summ_rows = []
        comp_map = {}
        if not segmentation_df.empty:
            comp_map = dict(zip(segmentation_df["Ticker"], segmentation_df["Composite"]))
            
        for t in backtest_tickers:
            s = sub_rets[t].dropna()
            sev = sub_rets.loc[severe_mask.values, t].dropna()
            
            avg_ret = float(s.mean()) if len(s) > 0 else 0.0
            med_ret = float(s.median()) if len(s) > 0 else 0.0
            hit_rate = float((s > 0).mean() * 100.0) if len(s) > 0 else 0.0
            avg_sev = float(sev.mean()) if len(sev) > 0 else np.nan
            hit_sev = float((sev > 0).mean() * 100.0) if len(sev) > 0 else np.nan
            best_day = float(s.max()) if len(s) > 0 else 0.0
            worst_day = float(s.min()) if len(s) > 0 else 0.0
            std_dev = float(s.std()) if len(s) > 1 else 0.0
            cum_ret = float(((1 + s / 100.0).prod() - 1.0) * 100.0) if len(s) > 0 else 0.0
            
            comp_val = comp_map.get(t, 0.0)
            if comp_val < 0.10:
                source_tag = "Sell-off Winner" if comp_val < -0.10 else "Chop Neutral"
            else:
                source_tag = "Trend Follower"
                
            ticker_info = info_dict.get(t, {"Strength": "—", "FDTS": "⚪ Neutral"})
            strength_val = ticker_info.get("Strength", "—")
            fdts_val = ticker_info.get("FDTS", "⚪ Neutral")
            
            summ_rows.append({
                "Ticker": t,
                "Source": source_tag,
                "FDTS": fdts_val,
                "Strength": strength_val,
                "Composite_Corr": round(comp_val, 4),
                "Selloff_Days": int(s.shape[0]),
                "Avg_Ret_%": round(avg_ret, 3),
                "Median_Ret_%": round(med_ret, 3),
                "Hit_Rate_Up_%": round(hit_rate, 1),
                "Avg_on_Severe_%": round(avg_sev, 3) if not np.isnan(avg_sev) else None,
                "HitRate_Severe_%": round(hit_sev, 1) if not np.isnan(hit_sev) else None,
                "Best_Day_%": round(best_day, 2),
                "Worst_Day_%": round(worst_day, 2),
                "Std_%": round(std_dev, 3),
                "Total_Cum_%": round(cum_ret, 2),
            })
            
        summary_df = pd.DataFrame(summ_rows)
        if not summary_df.empty:
            summary_df["Hedge_Score"] = (summary_df["Hit_Rate_Up_%"] * 0.5
                                      + summary_df["Avg_Ret_%"].clip(lower=-5.0) * 10.0
                                      + summary_df["HitRate_Severe_%"].fillna(0.0) * 0.3)
            summary_df = summary_df.sort_values(["Avg_Ret_%", "Hit_Rate_Up_%"], ascending=False).reset_index(drop=True)
            summary_df.insert(0, "Rank", range(1, len(summary_df) + 1))
            backtest_summary = summary_df
            
    return {
        "corr": corr,
        "segmentation": segmentation_df,
        "bucket_summary": bucket_summary,
        "backtest_summary": backtest_summary,
        "backtest_detail": backtest_detail,
    }


def generate_excel_report(universe_name, method, start_date, end_date, prices, results):
    wb = Workbook()
    
    HDR = PatternFill("solid", fgColor="1F4E79")
    ALT1 = PatternFill("solid", fgColor="DEEAF1")
    ALT2 = PatternFill("solid", fgColor="FFFFFF")
    GREEN = PatternFill("solid", fgColor="E2EFDA")
    GREEN2 = PatternFill("solid", fgColor="C6E0B4")
    SALMON = PatternFill("solid", fgColor="FCE4D6")
    RED = PatternFill("solid", fgColor="F8696B")
    WHITE_F = Font(color="FFFFFF", bold=True, size=11)
    BOLD = Font(bold=True)
    THIN = Side(style="thin", color="BFBFBF")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    def write_table(ws, df, start_row=1, corr_cols=None, bucket_col=None, ret_cols=None):
        corr_cols = corr_cols or []
        ret_cols = ret_cols or []
        cols = list(df.columns)
        
        for j, col in enumerate(cols, start=1):
            cell = ws.cell(row=start_row, column=j, value=col)
            cell.fill = HDR
            cell.font = WHITE_F
            cell.alignment = CENTER
            cell.border = BORDER
            
        for i, (_, r) in enumerate(df.iterrows(), start=start_row + 1):
            for j, col in enumerate(cols, start=1):
                val = r[col]
                if isinstance(val, (np.floating,)):
                    val = float(val)
                elif isinstance(val, (np.integer,)):
                    val = int(val)
                elif isinstance(val, float) and np.isnan(val):
                    val = None
                    
                cell = ws.cell(row=i, column=j, value=val)
                cell.border = BORDER
                cell.alignment = LEFT if isinstance(val, str) and len(str(val)) > 12 else CENTER
                cell.fill = ALT1 if (i % 2 == 0) else ALT2
                
                BUCKET_FILL = {
                    "1 - High-Beta Leaders": PatternFill("solid", fgColor="C6E0B4"),
                    "2 - Core Trend": PatternFill("solid", fgColor="E2EFDA"),
                    "3 - Low-Beta Followers": PatternFill("solid", fgColor="FFF2CC"),
                    "4 - Decoupled / Neutral": PatternFill("solid", fgColor="DEEAF1"),
                    "5 - Uncorrelated Hedge": PatternFill("solid", fgColor="D9D9D9"),
                    "6 - Inverse (Sell-off Winners)": PatternFill("solid", fgColor="FCE4D6"),
                }
                if bucket_col and col == bucket_col and val in BUCKET_FILL:
                    cell.fill = BUCKET_FILL[val]
                    
                if col in corr_cols and isinstance(val, (int, float)) and val is not None:
                    if val >= 0.5:
                        cell.fill = PatternFill("solid", fgColor="63BE7B")
                    elif val >= 0.25:
                        cell.fill = PatternFill("solid", fgColor="C6E0B4")
                    elif val > -0.1:
                        cell.fill = ALT1 if (i % 2 == 0) else ALT2
                    elif val > -0.3:
                        cell.fill = PatternFill("solid", fgColor="FCE4D6")
                    else:
                        cell.fill = PatternFill("solid", fgColor="F8696B")
                        
                if col in ret_cols and isinstance(val, (int, float)) and val is not None:
                    if val >= 1.0:
                        cell.fill = GREEN2
                    elif val > 0:
                        cell.fill = GREEN
                    elif val > -1.0:
                        cell.fill = SALMON
                    else:
                        cell.fill = RED
                        
                if col == "FDTS" and val:
                    if "Buy" in str(val):
                        cell.font = Font(color="22C55E", bold=True)
                    elif "Sell" in str(val):
                        cell.font = Font(color="EF4444", bold=True)
                if col == "Strength" and val:
                    if "▲" in str(val):
                        cell.font = Font(color="00D4AA", bold=True)
                    elif "▼" in str(val):
                        cell.font = Font(color="FF4B4B", bold=True)
                    elif "▶" in str(val):
                        cell.font = Font(color="FFA421", bold=True)
        ws.freeze_panes = ws.cell(row=start_row + 1, column=2)
        
    def autosize(ws, widths=None):
        widths = widths or {}
        for col_cells in ws.columns:
            first = col_cells[0]
            letter = get_column_letter(first.column)
            if letter in widths:
                ws.column_dimensions[letter].width = widths[letter]
                continue
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[letter].width = min(max(length + 2, 9), 42)

    # Sheet 1: Playbook / Methodology
    ws0 = wb.active
    ws0.title = "Playbook"
    ws0["A1"] = "Ticker Correlation Segmentation - Regime Shifting Engine"
    ws0["A1"].font = Font(bold=True, size=16, color="1F4E79")
    ws0.merge_cells("A1:H1")
    
    notes = [
        "",
        "STRATEGY CONTEXT: 25-delta calendar, long 40 DTE / short 20 DTE.",
        "Calendars are long vega + long theta -> they want the underlying to stay",
        "near the strike while front-month IV decays faster than back-month.",
        "Direction is expressed by strike placement + which underlying you choose.",
        "",
        "BENCHMARKS:  S&P 500 = SPY   |   NASDAQ = QQQ   |   RUSSELL = IWM",
        "",
        "SHIFTING MECHANISM - PRIORITY ORDER (highest first):",
        "  1) CORRELATION REGIME ALIGNMENT (primary) - sign & magnitude of a name's",
        "     correlation to the active benchmark decides which bucket is 'ON' and",
        "     whether to run CALL vs PUT calendars.",
        "  2) VOLATILITY (secondary) - realized vol vs IV and IV term-structure slope.",
        "     Low realized vol + positive (40DTE>20DTE) IV slope = best calendar.",
        "     Columns provisioned on 'Master Ranking' - populate from your vol feed.",
        "  3) OTHER (tertiary) - cross-benchmark consistency, idiosyncratic profile,",
        "     liquidity. Used only to break ties between names of equal correlation.",
        "",
        "REGIME PLAYBOOK:",
        "  BULL / TRENDING tape   -> Buckets 1-2 (High-Beta Leaders / Core Trend):",
        "      run CALL calendars, ride the trend. Highest Bull_Priority first.",
        "  SELL-OFF / RISK-OFF    -> Bucket 6 (Inverse): these RISE -> CALL calendars.",
        "      Buckets 1-2 fall -> PUT calendars or stand aside. Use Selloff_Priority.",
        "  CHOP / RANGE tape      -> Bucket 4 (Decoupled) + low-beta: delta-neutral",
        "      ATM calendars to harvest theta. Use Chop_Priority.",
        "",
        "BUCKET DEFINITIONS (by composite corr to SPY/QQQ/IWM):",
        "  1 High-Beta Leaders        composite >= 0.50",
        "  2 Core Trend               0.35 - 0.50",
        "  3 Low-Beta Followers       0.20 - 0.35",
        "  4 Decoupled / Neutral      0.05 - 0.20",
        "  5 Uncorrelated Hedge      -0.10 - 0.05",
        "  6 Inverse (Sell-off Win)   composite < -0.10",
        "",
        "NOTE: 'Consistent' = all three benchmarks agree in sign and spread <= 0.20,",
        "       i.e. a reliable directional play across SPY/QQQ/IWM.",
    ]
    for k, line in enumerate(notes, start=2):
        cell = ws0.cell(row=k, column=1, value=line)
        if line.endswith(":") or line.startswith("STRATEGY") or line.startswith("SHIFTING") \
                or line.startswith("REGIME") or line.startswith("BUCKET") or line.startswith("BENCHMARKS"):
            cell.font = BOLD
    ws0.column_dimensions["A"].width = 100
    
    # Sheet 2: Master Ranking
    ws1 = wb.create_sheet("Master Ranking")
    seg_df = results["segmentation"]
    master_cols = [
        "Rank", "Ticker", "Is_Reference", "Bucket", "Best_Proxy", "FDTS", "Strength",
        "Corr_SPY", "Corr_QQQ", "Corr_IWM", "Composite", "Bench_Spread", "Consistent",
        "Avg_Corr_Universe", "Idiosyncratic_Score", "RealizedVol_20d",
        "Bull_Priority", "Selloff_Priority", "Chop_Priority",
        "Bull_Action", "Selloff_Action", "Chop_Action"
    ]
    existing_cols = [c for c in master_cols if c in seg_df.columns]
    write_table(ws1, seg_df[existing_cols], corr_cols=["Corr_SPY", "Corr_QQQ", "Corr_IWM", "Composite"], bucket_col="Bucket")
    autosize(ws1, widths={"A": 7, "D": 26, "R": 40, "S": 40, "T": 40})
    
    # Sheet 3: Bull-Trend candidates
    ws2 = wb.create_sheet("Bull Trend (Ride)")
    bull_df = seg_df[seg_df["Composite"] >= 0.35].sort_values("Bull_Priority", ascending=False)
    bull_cols = ["Ticker", "Is_Reference", "Bucket", "Best_Proxy", "FDTS", "Strength", "Corr_SPY", "Corr_QQQ", "Corr_IWM",
                 "Composite", "Consistent", "Bull_Priority", "Bull_Action"]
    existing_bull = [c for c in bull_cols if c in bull_df.columns]
    write_table(ws2, bull_df[existing_bull], corr_cols=["Corr_SPY", "Corr_QQQ", "Corr_IWM", "Composite"], bucket_col="Bucket")
    autosize(ws2, widths={"C": 24, "K": 48})
    
    # Sheet 4: Sell-off winners
    ws3 = wb.create_sheet("Sell-off Winners")
    selloff_df = seg_df[seg_df["Composite"] < 0.10].sort_values("Selloff_Priority", ascending=False)
    sell_cols = ["Ticker", "Is_Reference", "Bucket", "Best_Proxy", "FDTS", "Strength", "Corr_SPY", "Corr_QQQ", "Corr_IWM",
                 "Composite", "Consistent", "Selloff_Priority", "Selloff_Action"]
    existing_sell = [c for c in sell_cols if c in selloff_df.columns]
    write_table(ws3, selloff_df[existing_sell], corr_cols=["Corr_SPY", "Corr_QQQ", "Corr_IWM", "Composite"], bucket_col="Bucket")
    autosize(ws3, widths={"C": 24, "K": 48})
    
    # Sheet 5: Chop Neutral
    ws4 = wb.create_sheet("Chop Neutral")
    chop_df = seg_df[(seg_df["Composite"] >= -0.10) & (seg_df["Composite"] < 0.25)].sort_values("Chop_Priority", ascending=False)
    chop_cols = ["Ticker", "Is_Reference", "Bucket", "Best_Proxy", "FDTS", "Strength", "Corr_SPY", "Corr_QQQ", "Corr_IWM",
                 "Composite", "Idiosyncratic_Score", "Chop_Priority", "Chop_Action"]
    existing_chop = [c for c in chop_cols if c in chop_df.columns]
    write_table(ws4, chop_df[existing_chop], corr_cols=["Corr_SPY", "Corr_QQQ", "Corr_IWM", "Composite"], bucket_col="Bucket")
    autosize(ws4, widths={"C": 24, "K": 48})
    
    # Sheet 6: Bucket Summary
    ws5 = wb.create_sheet("Bucket Summary")
    write_table(ws5, results["bucket_summary"], corr_cols=["Avg_Composite", "Avg_SPY", "Avg_QQQ", "Avg_IWM"], bucket_col="Bucket")
    autosize(ws5, widths={"A": 30})
    
    # Sheet 7: Regime Shift Map
    ws6 = wb.create_sheet("Regime Shift Map")
    shift = pd.DataFrame({
        "Active Regime": ["BULL / TRENDING", "SELL-OFF / RISK-OFF", "CHOP / RANGE"],
        "Detect Via": [
            "Benchmark above rising 20/50d MA; corr breadth high",
            "Benchmark below MA; breadth collapse; BONDS/GOLD corr turns positive",
            "Benchmark flat; falling realized vol; low trend strength",
        ],
        "Primary Bucket (P1 corr)": [
            "1 High-Beta Leaders + 2 Core Trend",
            "6 Inverse (Sell-off Winners)",
            "4 Decoupled / Neutral",
        ],
        "Calendar Side": [
            "CALL calendars (slightly OTM)",
            "CALL calendars on inverse names; PUT calendars on leaders",
            "ATM delta-neutral calendars",
        ],
        "Rank By": ["Bull_Priority", "Selloff_Priority", "Chop_Priority"],
        "P2 Vol Filter": [
            "Prefer low RV vs IV, positive IV term slope",
            "Watch IV spike - front-month rich can hurt long calendars",
            "Low realized vol ideal - max theta capture",
        ],
        "P3 Tie-break": [
            "Consistent=TRUE, higher liquidity",
            "Consistent=TRUE inverse, deepest negative composite",
            "Highest Idiosyncratic_Score",
        ],
    })
    write_table(ws6, shift)
    autosize(ws6, widths={"A": 22, "B": 46, "C": 30, "D": 40, "E": 16, "F": 40, "G": 34})
    
    # Backtest sheets
    if results.get("backtest_summary") is not None:
        ws_s = wb.create_sheet("Selloff Summary")
        summ_df = results["backtest_summary"]
        ret_cols_s = ["Avg_Ret_%", "Median_Ret_%", "Avg_on_Severe_%", "Best_Day_%", "Worst_Day_%", "Total_Cum_%"]
        write_table(ws_s, summ_df, corr_cols=["Composite_Corr"], ret_cols=ret_cols_s)
        autosize(ws_s)
        wb.move_sheet("Selloff Summary", -(len(wb.sheetnames) - wb.sheetnames.index("Selloff Summary") - 1))
        
    if results.get("backtest_detail") is not None:
        ws_d = wb.create_sheet("Selloff Detail")
        det_df = results["backtest_detail"]
        ret_cols_d = ["SPY_%", "QQQ_%", "IWM_%", "Mkt_Avg_%"] + [c for c in det_df.columns if c not in ["Date", "Severity", "Tickers_Up", "Pct_Tickers_Up"]]
        write_table(ws_d, det_df, ret_cols=ret_cols_d)
        autosize(ws_d)
        
    out_io = io.BytesIO()
    wb.save(out_io)
    out_io.seek(0)
    return out_io.getvalue()

# --------------------------------------------------------------------------- #
# Module Declaration
# --------------------------------------------------------------------------- #

class CorrelationMatrixModule(FazDaneModule):
    MODULE_NAME = "Correlation Matrix"
    MODULE_ICON = "🧮"
    MODULE_DESCRIPTION = "Multi-tab cross-asset return correlations, segmentation playbook, and index down-time backtest."
    TIER = 2
    SOURCE_NOTEBOOK = "Colab Correlation Matrix"
    CACHE_TTL = 3600
    REQUIRES_LIVE_DATA = True
    DATA_SOURCES = ["yfinance", "SQLite"]

    def render_sidebar(self):
        self._default_correlation_universe()
        st.markdown("**Correlation Universe**")
        self.universe_name, tickers, _ = render_universe_manager(
            key_prefix="corr",
            show_benchmark=False,
            label="Ticker Universe:",
        )
        self.tickers = tickers
        st.caption(f"{len(self.tickers)} assets selected from {self.universe_name}.")

        st.markdown("**Date Range**")
        today = datetime.today().date()
        default_start = today - pd.DateOffset(months=3)
        self.start_date = st.date_input("Start Date:", value=default_start.date(), key="corr_start")
        self.end_date = st.date_input("End Date:", value=today, key="corr_end")

        st.markdown("**Calculation**")
        self.method = st.selectbox("Correlation Method:", ["pearson", "spearman", "kendall"], index=0, key="corr_method")
        self.use_friendly_names = st.checkbox("Use friendly asset names", value=True, key="corr_friendly")

        st.markdown("**Data Storage & Runs**")
        self.latest_run = correlation_store.fetch_latest_run(self.universe_name)
        if self.latest_run:
            run_time = datetime.fromisoformat(self.latest_run["run_datetime"]).strftime("%Y-%m-%d %H:%M")
            st.info(f"📂 **Cached Run Found**\nSaved: {run_time}\nRange: {self.latest_run['start_date']} to {self.latest_run['end_date']}")
            
            self.run_source = st.radio(
                "Select Execution Source:",
                ["Use Cached SQLite Run", "Run Live calculations (yfinance)"],
                index=0,
                key="corr_run_source"
            )
        else:
            self.run_source = "Run Live calculations (yfinance)"
            st.caption("No cached runs found for this universe. Running live fetching is required.")

        st.divider()
        if self.run_source == "Run Live calculations (yfinance)":
            if st.button("Run Live Calculations", width="stretch", type="primary", key="corr_refresh"):
                fetch_close_prices.clear()
                st.rerun()

    def render_main(self):
        self.render_section_header(
            "🧮 Advanced Correlation Suite",
            "Cross-asset return correlations, bucketed playbook priorities, and market stress-test validation.",
        )

        if len(self.tickers) < 2:
            st.warning("Select at least two tickers for correlation analysis in the sidebar.")
            return

        if self.start_date >= self.end_date:
            st.warning("Start date must be before end date.")
            return

        prices = pd.DataFrame()
        results = {}
        loaded_from_db = False

        if self.run_source == "Use Cached SQLite Run" and self.latest_run is not None:
            prices = self.latest_run["prices_df"]
            results = self.latest_run["results"]
            loaded_from_db = True
        else:
            benchmarks = ["SPY", "QQQ", "IWM"]
            fetch_tickers = list(self.tickers)
            for b in benchmarks:
                if b not in fetch_tickers:
                    fetch_tickers.append(b)

            with st.spinner(f"Downloading price history from yfinance for {len(fetch_tickers)} assets..."):
                prices, info = fetch_close_prices(tuple(fetch_tickers), self.start_date, self.end_date)

            if prices.empty or prices.shape[1] < 2:
                st.warning("Not enough data was returned to compute a correlation matrix.")
                return

            results = compute_all_results(prices, self.tickers, self.method, info)
            
            try:
                correlation_store.save_run(
                    universe_name=self.universe_name,
                    tickers=self.tickers,
                    start_date=self.start_date.strftime("%Y-%m-%d"),
                    end_date=self.end_date.strftime("%Y-%m-%d"),
                    method=self.method,
                    prices_df=prices,
                    results_dict=results
                )
                st.toast("✅ Calculations saved to SQLite database successfully!")
            except Exception as e:
                st.sidebar.error(f"Failed to save run to SQLite: {e}")

        corr_matrix = results["corr"]
        display_tickers = [t for t in self.tickers if t in corr_matrix.columns]
        display_corr = corr_matrix.loc[display_tickers, display_tickers]

        if self.use_friendly_names:
            display_corr = display_corr.rename(columns=DEFAULT_RENAME_MAP, index=DEFAULT_RENAME_MAP)
        asset_labels = display_corr.columns.tolist()
        display_corr = display_corr.reindex(index=asset_labels, columns=asset_labels)

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Universe", self.universe_name)
        m2.metric("Assets Used", str(len(display_tickers)))
        m3.metric("Observations", str(max(len(prices.dropna(how="all")) - 1, 0)))
        
        status_text = "📁 Cached SQLite" if loaded_from_db else "🌐 Live yfinance"
        m4.metric("Data Source", status_text)

        if loaded_from_db:
            run_dt_formatted = datetime.fromisoformat(self.latest_run["run_datetime"]).strftime("%Y-%m-%d %H:%M")
            st.caption(f"💡 This data was loaded instantly from the SQLite database. Last computed on: **{run_dt_formatted}**.")
            
            # Check if indicators are missing or placeholders in the cached run
            seg_df = results.get("segmentation")
            if seg_df is not None and ("FDTS" not in seg_df.columns or "Strength" not in seg_df.columns or (seg_df["FDTS"] == "—").all()):
                st.warning("⚠️ **Missing Indicators in Cached Run:** This cached run does not contain Heikin-Ashi FDTS and Strength indicators. Please select **Run Live calculations (yfinance)** in the sidebar and click **Run Live Calculations** to compute and persist them for this universe.")

        report_bytes = generate_excel_report(self.universe_name, self.method, self.start_date, self.end_date, prices, results)
        st.download_button(
            label="📥 Download Complete Correlation & Backtest Report (.xlsx)",
            data=report_bytes,
            file_name=f"correlation_report_{self.universe_name.replace(' ', '_').lower()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
            type="primary"
        )
        
        st.divider()

        tab_corr, tab_seg, tab_backtest = st.tabs([
            "🧮 Correlation Matrix",
            "📊 Asset Segmentation",
            "📉 Down-Time Backtest"
        ])

        # ══════════════════════════════════════════════════════════════════
        # TAB 1: CORRELATION MATRIX
        # ══════════════════════════════════════════════════════════════════
        with tab_corr:
            st.markdown("### Interactive Heatmap")
            fig = go.Figure(
                data=go.Heatmap(
                    z=display_corr.values,
                    x=asset_labels,
                    y=asset_labels,
                    zmin=-1,
                    zmax=1,
                    zmid=0,
                    colorscale=[
                        [0.0, "#ef4444"],
                        [0.5, "#facc15"],
                        [1.0, "#22c55e"],
                    ],
                    text=(display_corr * 100).round(0).astype(int).astype(str).add("%").values,
                    texttemplate="%{text}",
                    hovertemplate="<b>%{y} vs %{x}</b><br>Correlation: %{customdata:.1f}%<extra></extra>",
                    customdata=(display_corr * 100).values,
                    colorbar=dict(title="Corr"),
                    xgap=1,
                    ygap=1,
                )
            )
            fig.update_traces(textfont=dict(size=14, color="#111827", family="Arial Black"))
            fig.update_layout(
                height=max(520, 42 * len(display_corr.index)),
                paper_bgcolor="#0d1b2e",
                plot_bgcolor="#0d1b2e",
                font=dict(color="#e2e8f0", size=14, family="Arial Black"),
                margin=dict(l=140, r=20, t=80, b=20),
                xaxis=dict(
                    side="top",
                    tickmode="array",
                    tickvals=asset_labels,
                    ticktext=asset_labels,
                    tickangle=-35,
                    tickfont=dict(size=13, family="Arial Black"),
                    automargin=True,
                ),
                yaxis=dict(
                    autorange="reversed",
                    tickmode="array",
                    tickvals=asset_labels,
                    ticktext=asset_labels,
                    tickfont=dict(size=13, family="Arial Black"),
                    automargin=True,
                ),
            )
            st.plotly_chart(fig, width="stretch", theme=None)

            st.markdown("### Correlation Table")
            st.dataframe(
                format_correlation_table(display_corr),
                width="stretch",
                height=min(760, 38 * (len(display_corr.index) + 1)),
            )

            st.download_button(
                "Download Correlations CSV",
                data=display_corr.to_csv(index=True),
                file_name=f"correlations_{self.universe_name.replace(' ', '_').lower()}.csv",
                mime="text/csv",
                width="stretch",
                key="corr_csv_dl"
            )

        # ══════════════════════════════════════════════════════════════════
        # TAB 2: ASSET SEGMENTATION
        # ══════════════════════════════════════════════════════════════════
        with tab_seg:
            seg_df = results["segmentation"]
            
            # Fallback for older saved runs
            for col in ["FDTS", "Strength"]:
                if col not in seg_df.columns:
                    seg_df[col] = "—"
            
            with st.expander("📋 View Regime Playbook Overview", expanded=False):
                playbook_grid = [
                    {"b": "1 - High-Beta Leaders", "bull": "PRIMARY - Call calendars (ride trend)", "bear": "AVOID / Put calendars - fall hardest", "chop": "ATM calendars - high theta"},
                    {"b": "2 - Core Trend", "bull": "Core - Call calendars ATM to slightly OTM", "bear": "Reduce / Put calendars", "chop": "ATM calendars"},
                    {"b": "3 - Low-Beta Followers", "bull": "Secondary - milder participation", "bear": "Hold - smaller drawdowns than leaders", "chop": "Good ATM candidates (lower whipsaw)"},
                    {"b": "4 - Decoupled / Neutral", "bull": "Neutral - stock-specific, size down", "bear": "Neutral - relative safe harbor", "chop": "PRIMARY - delta-neutral ATM calendars"},
                    {"b": "5 - Uncorrelated Hedge", "bull": "Diversifier - low trend capture", "bear": "Defensive ballast", "chop": "Neutral calendars / portfolio ballast"},
                    {"b": "6 - Inverse (Sell-off Winners)", "bull": "AVOID / Put calendars - drop on index rise", "bear": "PRIMARY - Call calendars (rise in risk-off)", "chop": "Tactical hedge overlay"}
                ]
                
                cols = st.columns(3)
                for idx, item in enumerate(playbook_grid):
                    with cols[idx % 3]:
                        st.markdown(
                            f"""
                            <div style='
                                background-color: #0e1e38; 
                                border: 1px solid #1f3d6b; 
                                border-radius: 8px; 
                                padding: 15px; 
                                margin-bottom: 15px;
                                height: 220px;
                            '>
                                <span style='font-size: 15px; font-weight: bold; color: #facc15;'>{item["b"]}</span>
                                <hr style='margin: 8px 0; border-color: #1f3d6b;'/>
                                <div style='font-size: 12px; line-height: 1.5;'>
                                    🟢 <b>Bull:</b> {item["bull"]}<br/>
                                    🔴 <b>Sell-off:</b> {item["bear"]}<br/>
                                    🟡 <b>Chop:</b> {item["chop"]}
                                </div>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )

            st.markdown("### 🏆 Priority Candidate Lists")
            sub1, sub2, sub3 = st.tabs([
                "🟢 Bull Trend Candidates",
                "🔴 Sell-off Hedges (Inverse)",
                "🟡 Chop Neutral Candidates"
            ])
            
            with sub1:
                st.caption("Composite Corr vs benchmarks >= 0.35, ranked by Bull_Priority descending (rewards high composite + cross-bench consistency)")
                bull_cands = seg_df[seg_df["Composite"] >= 0.35].sort_values("Bull_Priority", ascending=False)
                if not bull_cands.empty:
                    st.dataframe(
                        bull_cands[["Ticker", "Bucket", "Best_Proxy", "FDTS", "Strength", "Corr_SPY", "Corr_QQQ", "Corr_IWM", "Composite", "Consistent", "RealizedVol_20d", "Bull_Priority", "Bull_Action"]]
                        .style.map(color_correlation, subset=["Corr_SPY", "Corr_QQQ", "Corr_IWM", "Composite"])
                        .map(color_bucket, subset=["Bucket"])
                        .map(style_fdts, subset=["FDTS"])
                        .map(style_strength, subset=["Strength"])
                        .format({"Corr_SPY": "{:.2f}", "Corr_QQQ": "{:.2f}", "Corr_IWM": "{:.2f}", "Composite": "{:.2f}", "Bull_Priority": "{:.1f}", "RealizedVol_20d": "{:.1f}%"}),
                        width="stretch",
                        hide_index=True
                    )
                else:
                    st.info("No candidates match this criteria.")
                    
            with sub2:
                st.caption("Composite Corr vs benchmarks < 0.10, ranked by Selloff_Priority descending (rewards strongly negative composite correlation)")
                selloff_cands = seg_df[seg_df["Composite"] < 0.10].sort_values("Selloff_Priority", ascending=False)
                if not selloff_cands.empty:
                    st.dataframe(
                        selloff_cands[["Ticker", "Bucket", "Best_Proxy", "FDTS", "Strength", "Corr_SPY", "Corr_QQQ", "Corr_IWM", "Composite", "Consistent", "RealizedVol_20d", "Selloff_Priority", "Selloff_Action"]]
                        .style.map(color_correlation, subset=["Corr_SPY", "Corr_QQQ", "Corr_IWM", "Composite"])
                        .map(color_bucket, subset=["Bucket"])
                        .map(style_fdts, subset=["FDTS"])
                        .map(style_strength, subset=["Strength"])
                        .format({"Corr_SPY": "{:.2f}", "Corr_QQQ": "{:.2f}", "Corr_IWM": "{:.2f}", "Composite": "{:.2f}", "Selloff_Priority": "{:.1f}", "RealizedVol_20d": "{:.1f}%"}),
                        width="stretch",
                        hide_index=True
                    )
                else:
                    st.info("No candidates match this criteria.")
                    
            with sub3:
                st.caption("Composite Corr between -0.10 and 0.25, ranked by Chop_Priority descending (rewards near-zero composite + high idiosyncratic independence)")
                chop_cands = seg_df[(seg_df["Composite"] >= -0.10) & (seg_df["Composite"] < 0.25)].sort_values("Chop_Priority", ascending=False)
                if not chop_cands.empty:
                    st.dataframe(
                        chop_cands[["Ticker", "Bucket", "Best_Proxy", "FDTS", "Strength", "Corr_SPY", "Corr_QQQ", "Corr_IWM", "Composite", "Idiosyncratic_Score", "RealizedVol_20d", "Chop_Priority", "Chop_Action"]]
                        .style.map(color_correlation, subset=["Corr_SPY", "Corr_QQQ", "Corr_IWM", "Composite"])
                        .map(color_bucket, subset=["Bucket"])
                        .map(style_fdts, subset=["FDTS"])
                        .map(style_strength, subset=["Strength"])
                        .format({"Corr_SPY": "{:.2f}", "Corr_QQQ": "{:.2f}", "Corr_IWM": "{:.2f}", "Composite": "{:.2f}", "Idiosyncratic_Score": "{:.2f}", "Chop_Priority": "{:.1f}", "RealizedVol_20d": "{:.1f}%"}),
                        width="stretch",
                        hide_index=True
                    )
                else:
                    st.info("No candidates match this criteria.")

            st.markdown("### 👑 Master Segmentation Ranking")
            st.dataframe(
                seg_df
                .style.map(color_correlation, subset=["Corr_SPY", "Corr_QQQ", "Corr_IWM", "Composite"])
                .map(color_bucket, subset=["Bucket"])
                .map(style_fdts, subset=["FDTS"])
                .map(style_strength, subset=["Strength"])
                .format({
                    "Corr_SPY": "{:.2f}", "Corr_QQQ": "{:.2f}", "Corr_IWM": "{:.2f}",
                    "Composite": "{:.2f}", "Min_Corr": "{:.2f}", "Max_Corr": "{:.2f}",
                    "Bench_Spread": "{:.2f}", "Avg_Corr_Universe": "{:.2f}", "Idiosyncratic_Score": "{:.2f}",
                    "RealizedVol_20d": "{:.1f}%", "Bull_Priority": "{:.1f}", "Selloff_Priority": "{:.1f}",
                    "Chop_Priority": "{:.1f}"
                }),
                width="stretch",
                hide_index=True
            )

            st.markdown("### 📊 Bucket Summary Analytics")
            st.dataframe(
                results["bucket_summary"]
                .style.map(color_correlation, subset=["Avg_Composite", "Avg_SPY", "Avg_QQQ", "Avg_IWM"])
                .map(color_bucket, subset=["Bucket"])
                .format({"Avg_Composite": "{:.3f}", "Avg_SPY": "{:.3f}", "Avg_QQQ": "{:.3f}", "Avg_IWM": "{:.3f}"}),
                width="stretch",
                hide_index=True
            )

        # ══════════════════════════════════════════════════════════════════
        # TAB 3: DOWN-TIME BACKTEST
        # ══════════════════════════════════════════════════════════════════
        with tab_backtest:
            lookback_days = (self.end_date - self.start_date).days
            if lookback_days < 180:
                st.warning(
                    f"⚠️ **Short Lookback Warning ({lookback_days} days):** The selected date range is relatively short. "
                    "For a reliable down-time statistical backtest, a lookback of at least 6 months (preferably 1 year) "
                    "is recommended. Adjust 'Date Range' in the sidebar if needed."
                )

            backtest_summary = results["backtest_summary"]
            backtest_detail = results["backtest_detail"]

            # Fallback for older saved runs
            if backtest_summary is not None and not backtest_summary.empty:
                for col in ["FDTS", "Strength"]:
                    if col not in backtest_summary.columns:
                        backtest_summary[col] = "—"

            if backtest_summary is None or backtest_summary.empty:
                st.info("No index sell-off days were identified in this period (days where SPY, QQQ, and IWM all closed down).")
            else:
                total_trading_days = max(len(prices.dropna(how="all")) - 1, 0)
                n_selloff = len(backtest_detail)
                pct_selloff = (n_selloff / total_trading_days * 100.0) if total_trading_days > 0 else 0.0
                
                avg_mkt_ret = backtest_detail["Mkt_Avg_%"].mean()
                best_hedge_row = backtest_summary.sort_values("Hedge_Score", ascending=False).iloc[0]
                best_hedge_name = f"{best_hedge_row['Ticker']} (Score: {best_hedge_row['Hedge_Score']:.1f})"

                kpi1, kpi2, kpi3, kpi4 = st.columns(4)
                kpi1.metric("Trading Days", str(total_trading_days))
                kpi2.metric("Sell-off Days", f"{n_selloff} ({pct_selloff:.1f}%)")
                kpi3.metric("Avg Index Drop", f"{avg_mkt_ret:.2f}%")
                kpi4.metric("Top Hedge (Hedge Score)", best_hedge_name)

                st.markdown("### 📈 Down-Time Asset Performance")
                chart_df = backtest_summary.sort_values("Avg_Ret_%", ascending=False).copy()
                chart_df["Color"] = chart_df["Avg_Ret_%"].apply(lambda x: "#22c55e" if x >= 0 else "#ef4444")
                
                fig_bar = go.Figure(
                    data=go.Bar(
                        x=chart_df["Ticker"],
                        y=chart_df["Avg_Ret_%"],
                        marker_color=chart_df["Color"],
                        hovertemplate="<b>%{x}</b><br>Avg Return on Down Days: %{y:.2f}%<extra></extra>"
                    )
                )
                fig_bar.update_layout(
                    title="Average Daily Return on Index Sell-off Days",
                    xaxis_title="Ticker",
                    yaxis_title="Average Return (%)",
                    paper_bgcolor="#0d1b2e",
                    plot_bgcolor="#0d1b2e",
                    font=dict(color="#e2e8f0", size=13),
                    margin=dict(l=40, r=20, t=50, b=45),
                    xaxis=dict(tickangle=-45)
                )
                st.plotly_chart(fig_bar, width="stretch", theme=None)

                st.markdown("### 🏆 Hedge Performance Summary")
                st.dataframe(
                    backtest_summary
                    .style.map(color_return, subset=["Avg_Ret_%", "Median_Ret_%", "Avg_on_Severe_%", "Best_Day_%", "Worst_Day_%", "Total_Cum_%"])
                    .map(style_fdts, subset=["FDTS"])
                    .map(style_strength, subset=["Strength"])
                    .format({
                        "Composite_Corr": "{:.2f}", "Avg_Ret_%": "{:.3f}%", "Median_Ret_%": "{:.3f}%",
                        "Hit_Rate_Up_%": "{:.1f}%", "Avg_on_Severe_%": "{:.3f}%", "HitRate_Severe_%": "{:.1f}%",
                        "Best_Day_%": "{:.2f}%", "Worst_Day_%": "{:.2f}%", "Std_%": "{:.3f}%",
                        "Total_Cum_%": "{:.2f}%", "Hedge_Score": "{:.1f}"
                    }),
                    width="stretch",
                    hide_index=True
                )

                st.markdown("### 📅 Sell-off Day Log")
                st.dataframe(
                    backtest_detail
                    .style.map(color_return, subset=["SPY_%", "QQQ_%", "IWM_%", "Mkt_Avg_%"] + [c for c in backtest_summary["Ticker"].tolist() if c in backtest_detail.columns])
                    .format({
                        "SPY_%": "{:.2f}%", "QQQ_%": "{:.2f}%", "IWM_%": "{:.2f}%",
                        "Mkt_Avg_%": "{:.2f}%", "Pct_Tickers_Up": "{:.1f}%"
                    }),
                    width="stretch",
                    hide_index=True
                )

    def _default_correlation_universe(self):
        key = "corr_sel"
        target = "Correlation Matrix Assets"
        names = get_universe_names()
        if target in names and key not in st.session_state:
            st.session_state[key] = target
